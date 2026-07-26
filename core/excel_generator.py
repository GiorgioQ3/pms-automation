import os
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Generatore di Report Excel conforme al protocollo DPR-385 PSUR Worksheet."""

    SOURCES_KEYS = [
        "Minister of Health",
        "MAUDE",
        "BfArM",
        "MHRA",
        "Safety Communication (FDA)",
        "MD Recalls (FDA)",
        "Letters to Health Care Providers (FDA)",
        "National Vulnerability Database (NVD)"
    ]

    def __init__(self, file_path: str = "PMS_Report_DPR-385.xlsx"):
        self.file_path = file_path

    def generate(
        self,
        records: List[Dict[str, Any]],
        target_device: str = "Medical Device",
        search_period: str = "All Available Data",
        keywords_list: Optional[List[str]] = None
    ) -> str:
        wb = openpyxl.Workbook()
        
        # 1. Foglio Frontpage
        ws_front = wb.active
        ws_front.title = "Frontpage"
        self._build_frontpage(ws_front, target_device, search_period)

        # 2. Foglio PSUR_Summary
        ws_summary = wb.create_sheet(title="PSUR_Summary")
        self._build_summary(ws_summary, records, keywords_list or [target_device])

        # 3. Foglio Dettaglio_Incidenti
        ws_details = wb.create_sheet(title="Dettaglio_Incidenti")
        self._build_details(ws_details, records)

        wb.save(self.file_path)
        logger.info(f"Report DPR-385 salvato con successo in {self.file_path}")
        return self.file_path

    def generate_report(
        self,
        keyword_stats: Dict[str, Dict[str, Dict[str, int]]],
        all_records: List[Dict[str, Any]],
        output_filename: str,
        keyword_input_str: str,
        start_date: str,
        end_date: str,
        search_date: str
    ) -> str:
        self.file_path = output_filename
        wb = openpyxl.Workbook()

        # 1. Frontpage
        ws_front = wb.active
        ws_front.title = "Frontpage"
        self._build_frontpage(ws_front, target_device=keyword_input_str, search_period=f"{start_date} to {end_date}")

        # 2. PSUR_Summary
        ws_summary = wb.create_sheet(title="PSUR_Summary")
        self._build_custom_summary(ws_summary, keyword_stats)

        # 3. Dettaglio_Incidenti
        ws_details = wb.create_sheet(title="Dettaglio_Incidenti")
        self._build_details(ws_details, all_records)

        wb.save(output_filename)
        logger.info(f"Report DPR-385 salvato con successo in {output_filename}")
        return output_filename

    def _build_frontpage(self, ws, target_device: str, search_period: str):
        ws.views.sheetView[0].showGridLines = True
        
        ws['A2'] = "Title"
        ws['B2'] = "DPR-385 PSUR Worksheet"
        ws['A3'] = "Scope"
        ws['B3'] = f"This document serves as a worksheet for collecting vigilance data from the main DB for the PMS of {target_device}."
        ws['A4'] = "Search Period"
        ws['B4'] = search_period
        ws['A5'] = "Protocol"
        ws['B5'] = "DPR-385"
        ws['A6'] = "Version"
        ws['B6'] = "1.0"
        ws['A7'] = "Last update"
        ws['B7'] = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws['A8'] = "Label"
        ws['B8'] = "Internal"

        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for r in range(2, 9):
            ws.cell(row=r, column=1).fill = header_fill
            ws.cell(row=r, column=1).font = header_font
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(row=r, column=2).font = Font(name="Calibri", size=11, bold=(r==2))

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 80

    def _build_custom_summary(self, ws, keyword_stats: Dict[str, Dict[str, Dict[str, int]]]):
        ws.views.sheetView[0].showGridLines = True
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        ws.cell(row=1, column=1, value="ID")
        ws.cell(row=1, column=2, value="Keywords")
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).font = header_font
        ws.cell(row=1, column=2).fill = header_fill
        ws.cell(row=1, column=2).font = header_font

        col_idx = 3
        for src in self.SOURCES_KEYS:
            ws.cell(row=1, column=col_idx, value=src)
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+1)
            ws.cell(row=1, column=col_idx).fill = header_fill
            ws.cell(row=1, column=col_idx).font = header_font
            ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
            col_idx += 2

        current_row = 2
        for idx, (kw, stats) in enumerate(keyword_stats.items(), start=1):
            ws.cell(row=current_row, column=1, value=idx)
            ws.cell(row=current_row, column=2, value=f'"{kw}"')
            
            c_offset = 3
            for src in self.SOURCES_KEYS:
                src_data = stats.get(src, {"tot": 0, "dupl": 0, "sel": 0})
                ws.cell(row=current_row, column=c_offset, value="Tot:")
                ws.cell(row=current_row, column=c_offset+1, value=src_data.get("tot", 0))
                
                ws.cell(row=current_row+1, column=c_offset, value="Dupl:")
                ws.cell(row=current_row+1, column=c_offset+1, value=src_data.get("dupl", 0))
                
                ws.cell(row=current_row+2, column=c_offset, value="Sel:")
                ws.cell(row=current_row+2, column=c_offset+1, value=src_data.get("sel", 0))

                c_offset += 2

            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+2, end_column=1)
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row+2, end_column=2)
            current_row += 3

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 30

    def _build_summary(self, ws, records: List[Dict[str, Any]], keywords_list: List[str]):
        ws.views.sheetView[0].showGridLines = True

        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        ws.cell(row=1, column=1, value="ID")
        ws.cell(row=1, column=2, value="Keywords")
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).font = header_font
        ws.cell(row=1, column=2).fill = header_fill
        ws.cell(row=1, column=2).font = header_font

        col_idx = 3
        for src in self.SOURCES_KEYS:
            ws.cell(row=1, column=col_idx, value=src)
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+1)
            ws.cell(row=1, column=col_idx).fill = header_fill
            ws.cell(row=1, column=col_idx).font = header_font
            ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
            col_idx += 2

        current_row = 2
        for idx, kw in enumerate(keywords_list, start=1):
            ws.cell(row=current_row, column=1, value=idx)
            ws.cell(row=current_row, column=2, value=f'"{kw}"')
            
            if len(keywords_list) > 1:
                kw_recs = [
                    r for r in records
                    if r.get("searched_keyword", r.get("keyword", kw)).lower() == kw.lower()
                ]
            else:
                kw_recs = records

            c_offset = 3
            for src in self.SOURCES_KEYS:
                src_recs = [r for r in kw_recs if self._match_source(r.get("fonte", ""), src)]
                tot_count = len(src_recs)
                dupl_count = sum(1 for r in src_recs if r.get("is_duplicate", False))
                sel_count = tot_count - dupl_count
                
                ws.cell(row=current_row, column=c_offset, value="Tot:")
                ws.cell(row=current_row, column=c_offset+1, value=tot_count)
                
                ws.cell(row=current_row+1, column=c_offset, value="Dupl:")
                ws.cell(row=current_row+1, column=c_offset+1, value=dupl_count)
                
                ws.cell(row=current_row+2, column=c_offset, value="Sel:")
                ws.cell(row=current_row+2, column=c_offset+1, value=sel_count)

                c_offset += 2

            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+2, end_column=1)
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row+2, end_column=2)
            current_row += 3

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 30

    def _match_source(self, record_fonte: str, target_src: str) -> bool:
        f = record_fonte.lower()
        t = target_src.lower()
        if "minister" in t and ("ministero" in f or "health" in f): return True
        if "maude" in t and "maude" in f: return True
        if "bfarm" in t and "bfarm" in f: return True
        if "mhra" in t and "mhra" in f: return True
        if "safety communication" in t and "safety communication" in f: return True
        if "md recalls" in t and ("recalls" in f or "richiami" in f): return True
        if "letters to health" in t and "letter" in f: return True
        if "vulnerability" in t and ("nvd" in f or "cybersecurity" in f or "cve" in f): return True
        return False

    def _build_details(self, ws, records: List[Dict[str, Any]]):
        ws.views.sheetView[0].showGridLines = True
        headers = [
            "Fonte", "ID Segnalazione", "Data Pubblicazione", "Fabbricante",
            "Dispositivo", "Descrizione Evento", "Tipologia", "Tag Competitor", "URL Fonte"
        ]
        
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        link_font = Font(name="Calibri", size=11, color="0000FF", underline="single")

        for r_idx, rec in enumerate(records, start=2):
            ws.cell(row=r_idx, column=1, value=rec.get("fonte") or rec.get("source_name") or rec.get("source_site", ""))
            ws.cell(row=r_idx, column=2, value=rec.get("id_segnalazione") or rec.get("id") or rec.get("reference", ""))
            ws.cell(row=r_idx, column=3, value=rec.get("data_pubblicazione") or rec.get("data") or rec.get("date", ""))
            ws.cell(row=r_idx, column=4, value=rec.get("fabbricante", ""))
            ws.cell(row=r_idx, column=5, value=rec.get("dispositivo") or rec.get("titolo") or rec.get("title", ""))
            ws.cell(row=r_idx, column=6, value=rec.get("descrizione_evento") or rec.get("summary") or rec.get("titolo") or rec.get("title", ""))
            ws.cell(row=r_idx, column=7, value=rec.get("tipologia") or rec.get("tag", ""))
            ws.cell(row=r_idx, column=8, value=rec.get("tag_competitor") or rec.get("tag", "N/A"))
            
            url_val = rec.get("url_fonte") or rec.get("url") or rec.get("link_documento") or rec.get("link") or ""
            cell_url = ws.cell(row=r_idx, column=9, value=url_val)
            if str(url_val).startswith(("http://", "https://")):
                cell_url.hyperlink = str(url_val)
                cell_url.font = link_font

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)


def genera_excel_report(dati: List[Dict[str, Any]], output_path: str = "PMS_Report_DPR-385.xlsx") -> str:
    generator = ExcelGenerator(file_path=output_path)
    return generator.generate(records=dati)


def genera_excel_con_ipertesti(records: List[Dict[str, Any]], output_filename: str = "PMS_Report_Con_Link.xlsx") -> str:
    generator = ExcelGenerator(file_path=output_filename)
    return generator.generate(records=records)


def genera_tabella_riassuntiva_siti(raw_results: List[Dict[str, Any]]) -> Any:
    import pandas as pd
    siti_mappa = {}
    for record in raw_results:
        sito = record.get("source_site") or record.get("fonte") or record.get("source_name") or "Sconosciuto"
        if sito not in siti_mappa:
            siti_mappa[sito] = {"Sito / Fonte": sito, "Record Trovati": 0, "Stato Scansione": "Completato"}
        siti_mappa[sito]["Record Trovati"] += 1

    df_summary = pd.DataFrame(list(siti_mappa.values()))
    if df_summary.empty:
        df_summary = pd.DataFrame(columns=["Sito / Fonte", "Record Trovati", "Stato Scansione"])
    return df_summary

