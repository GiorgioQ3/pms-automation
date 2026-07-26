"""
Tools Module: preview_export_tool.py
Modulo per l'esportazione Excel con ipertesti formattati e l'anteprima GUI dei risultati.
"""

import openpyxl
from openpyxl.styles import Font
import tkinter as tk
from tkinter import ttk


# ==========================================
# 1. STRUTTURA DATI E GENERAZIONE EXCEL
# ==========================================

def esporta_excel_con_link(records: list, output_filename: str = "report_fsn.xlsx"):
    """
    Genera un file Excel inserendo per ogni record un ipertesto cliccabile.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Segnalazioni FSN"

    # Intestazioni della tabella Excel
    headers = ["Sito Fonte", "Titolo / Descrizione", "Codice Riferimento", "Link Record"]
    ws.append(headers)

    # Formattazione per la riga di intestazione (grassetto)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    # Popolamento dei dati
    for record in records:
        ws.append([
            record.get("site", ""),
            record.get("title", ""),
            record.get("ref", ""),
            "Apri Scheda / PDF"  # Testo visibile della cella
        ])

        # Recupera la riga appena aggiunta
        current_row = ws.max_row
        link_cell = ws.cell(row=current_row, column=4)

        # Assegna l'URL ipertestuale e lo stile grafico (blu e sottolineato)
        link_cell.hyperlink = record.get("url", "")
        link_cell.font = Font(color="0000FF", underline="single")

    # Autoadatta la larghezza delle colonne
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(output_filename)
    print(f"File salvato con successo: {output_filename}")
    return output_filename


# ==========================================
# 2. TABELLA RIASSUNTIVA NELLA GUI (Tkinter)
# ==========================================

def mostra_gui_anteprima(summary_data: list, raw_records: list):
    """
    Mostra una finestra con la tabella riassuntiva dei risultati 
    prima di procedere al download/salvataggio del file Excel.
    """
    root = tk.Tk()
    root.title("Risultati Ricerca Segnalazioni")
    root.geometry("500x350")

    # Etichetta di intestazione
    label = ttk.Label(root, text="Riepilogo Ricerca per Sito", font=("Arial", 12, "bold"))
    label.pack(pady=10)

    # Creazione della Tabella GUI (Treeview)
    cols = ("Sito", "Record Trovati", "Stato")
    tree = ttk.Treeview(root, columns=cols, show="headings", height=8)

    for col in cols:
        tree.heading(col, text=col)
        tree.column(col, anchor="center", width=140)

    # Inserimento dei dati riassuntivi nella tabella GUI
    for row in summary_data:
        tree.insert("", "end", values=(row["site"], row["count"], row["status"]))

    tree.pack(padx=10, pady=10, fill="both", expand=True)

    # Bottone per avviare il download / salvataggio Excel
    def aziona_download():
        esporta_excel_con_link(raw_records)
        root.destroy()

    btn_download = ttk.Button(root, text="Scarica Report Excel", command=aziona_download)
    btn_download.pack(pady=15)

    root.mainloop()


# ==========================================
# ESEMPIO DI ESECUZIONE / INTEGRAZIONE
# ==========================================
if __name__ == "__main__":
    summary_results = [
        {"site": "BfArM (Germania)", "count": 1, "status": "Completato"},
        {"site": "GOV.UK (Regno Unito)", "count": 0, "status": "Completato"},
        {"site": "FDA (USA)", "count": 2, "status": "Completato"}
    ]

    detailed_records = [
        {
            "site": "BfArM",
            "title": "Urgent Field Safety Notice for INFINITT PACS 7.0",
            "ref": "14094/26",
            "url": "https://www.bfarm.de/SharedDocs/Risikoinformationen/MDE/DE/14/14094-26.html"
        },
        {
            "site": "FDA",
            "title": "Medical Device Recall Notice Sample 1",
            "ref": "Z-0001-2026",
            "url": "https://www.fda.gov/medical-devices/recalls/sample1"
        },
        {
            "site": "FDA",
            "title": "Medical Device Recall Notice Sample 2",
            "ref": "Z-0002-2026",
            "url": "https://www.fda.gov/medical-devices/recalls/sample2"
        }
    ]

    mostra_gui_anteprima(summary_results, detailed_records)
