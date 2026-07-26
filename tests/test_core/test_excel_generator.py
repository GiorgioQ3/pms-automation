import pytest
from pathlib import Path
import openpyxl
from core.excel_generator import ExcelGenerator, genera_excel_report


def test_excel_generator_dpr385_structure(tmp_path: Path):
    output_file = tmp_path / "report_dpr385_test.xlsx"
    records = [
        {
            "fonte": "Safety Communication (FDA)",
            "id_segnalazione": "FDA-001",
            "data_pubblicazione": "15/03/2024",
            "fabbricante": "ACME Med",
            "dispositivo": "Mammography AI",
            "descrizione_evento": "Safety alert details",
            "tipologia": "Safety Communication",
            "tag_competitor": "Competitor A",
            "url_fonte": "https://example.com/fda1"
        },
        {
            "fonte": "Letters to Health Care Providers (FDA)",
            "id_segnalazione": "FDA-002",
            "data_pubblicazione": "10/01/2024",
            "fabbricante": "Siemens",
            "dispositivo": "CAD Software",
            "descrizione_evento": "Letter details",
            "tipologia": "Letter to Health Care Providers",
            "tag_competitor": "Siemens",
            "url_fonte": "https://example.com/fda2"
        }
    ]

    generator = ExcelGenerator(file_path=str(output_file))
    res_path = generator.generate(
        records=records,
        target_device="mammography",
        search_period="01/01/2024 to 31/12/2024",
        keywords_list=["mammography"]
    )
    assert Path(res_path).exists()

    wb = openpyxl.load_workbook(output_file)
    assert "Frontpage" in wb.sheetnames
    assert "PSUR_Summary" in wb.sheetnames
    assert "Dettaglio_Incidenti" in wb.sheetnames

    # Check Frontpage
    ws_front = wb["Frontpage"]
    assert ws_front["B2"].value == "DPR-385 PSUR Worksheet"
    assert ws_front["B4"].value == "01/01/2024 to 31/12/2024"

    # Check PSUR_Summary
    ws_sum = wb["PSUR_Summary"]
    assert ws_sum.cell(row=1, column=1).value == "ID"
    assert ws_sum.cell(row=1, column=2).value == "Keywords"

    # Check Dettaglio_Incidenti
    ws_det = wb["Dettaglio_Incidenti"]
    headers = [ws_det.cell(row=1, column=c).value for c in range(1, 10)]
    assert headers == [
        "Fonte", "ID Segnalazione", "Data Pubblicazione", "Fabbricante",
        "Dispositivo", "Descrizione Evento", "Tipologia", "Tag Competitor", "URL Fonte"
    ]
    assert ws_det.cell(row=2, column=1).value == "Safety Communication (FDA)"
    assert ws_det.cell(row=2, column=9).hyperlink is not None


def test_genera_excel_report_dati_vuoti(tmp_path: Path):
    output_file = tmp_path / "report_vuoto.xlsx"
    dati = []

    res_path = genera_excel_report(dati, str(output_file))
    assert Path(res_path).exists()

    wb = openpyxl.load_workbook(output_file)
    assert "Frontpage" in wb.sheetnames
    assert "PSUR_Summary" in wb.sheetnames
    assert "Dettaglio_Incidenti" in wb.sheetnames
