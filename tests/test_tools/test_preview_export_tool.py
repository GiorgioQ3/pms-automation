import openpyxl
from pathlib import Path
from tools.preview_export_tool import esporta_excel_con_link


def test_esporta_excel_con_link(tmp_path: Path):
    output_file = tmp_path / "test_report_fsn.xlsx"
    records = [
        {
            "site": "BfArM",
            "title": "Sample Title",
            "ref": "12345",
            "url": "https://example.com/test"
        }
    ]

    res_path = esporta_excel_con_link(records, str(output_file))
    assert Path(res_path).exists()

    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    assert ws.title == "Segnalazioni FSN"
    assert ws.cell(row=1, column=1).value == "Sito Fonte"
    assert ws.cell(row=2, column=4).value == "Apri Scheda / PDF"
    assert ws.cell(row=2, column=4).hyperlink.target == "https://example.com/test"
