import os
import openpyxl
import pandas as pd
from core.excel_generator import (
    genera_tabella_riassuntiva_siti,
    genera_excel_con_ipertesti,
    ExcelGenerator
)
from main import PMSOrchestrator


def test_genera_tabella_riassuntiva_siti():
    sample_records = [
        {
            "source_site": "BfArM (Germania)",
            "title": "Urgent Field Safety Notice",
            "reference": "14094/26",
            "url": "https://www.bfarm.de/test.html"
        },
        {
            "source_site": "BfArM (Germania)",
            "title": "Field Safety Notice 2",
            "reference": "14095/26",
            "url": "https://www.bfarm.de/test2.html"
        },
        {
            "source_site": "Salute.gov.it",
            "title": "Avviso di sicurezza",
            "reference": "2026-001",
            "url": "https://www.salute.gov.it/test.html"
        }
    ]

    df = genera_tabella_riassuntiva_siti(sample_records)
    assert isinstance(df, pd.DataFrame)
    assert not df.empty
    assert "Sito / Fonte" in df.columns
    assert "Record Trovati" in df.columns
    assert "Stato Scansione" in df.columns

    bfarm_row = df[df["Sito / Fonte"] == "BfArM (Germania)"]
    assert len(bfarm_row) == 1
    assert bfarm_row.iloc[0]["Record Trovati"] == 2


def test_genera_excel_con_ipertesti(tmp_path):
    output_file = str(tmp_path / "test_report.xlsx")
    sample_records = [
        {
            "fonte": "BfArM (Germania)",
            "id_segnalazione": "BfArM-001",
            "data_pubblicazione": "2026-01-15",
            "fabbricante": "Test Firm",
            "dispositivo": "Test Device",
            "descrizione_evento": "Safety alert details",
            "tipologia": "FSN",
            "tag_competitor": "N/A",
            "url": "https://www.bfarm.de/notice/1"
        }
    ]

    result_path = genera_excel_con_ipertesti(sample_records, output_filename=output_file)
    assert os.path.exists(result_path)

    wb = openpyxl.load_workbook(result_path)
    assert "Dettaglio_Incidenti" in wb.sheetnames

    ws = wb["Dettaglio_Incidenti"]
    cell_url = ws.cell(row=2, column=9)
    assert cell_url.value == "https://www.bfarm.de/notice/1"
    assert cell_url.hyperlink is not None
    assert cell_url.hyperlink.target == "https://www.bfarm.de/notice/1"


def test_pms_orchestrator_source_summary():
    orchestrator = PMSOrchestrator()
    res = orchestrator.run_pipeline("mammography", start_date="2026-01-01", end_date="2026-07-26")

    assert "source_summary" in res
    summary = res["source_summary"]
    assert isinstance(summary, list)
    assert len(summary) == 8

    for item in summary:
        assert "Sito / Fonte" in item
        assert "Record Trovati" in item
        assert "Record Selezionati" in item
        assert "Stato Scansione" in item
        assert item["Stato Scansione"] in ["Completato", "0 Risultati", "Errore"]
